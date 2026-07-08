# -*- coding: utf-8 -*-
"""
事件导出与 Excel 转换工具
"""
import os
import json
import shutil
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from models.data import Event, VehicleInfo
from utils.helpers import ensure_dir


class EventExporter:
    """事件导出器"""

    def __init__(self, output_dir: str):
        self.output_dir = output_dir

    def export_events(self, events: List[Event], vehicle_info: VehicleInfo) -> str:
        """导出事件到本地目录，返回导出目录"""
        export_dir = os.path.join(
            self.output_dir,
            f"DSSAD_EXPORT_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )
        ensure_dir(export_dir)

        # 写入车辆信息
        with open(os.path.join(export_dir, "vehicle_info.json"), "w", encoding="utf-8") as f:
            json.dump({
                "vin": vehicle_info.vin,
                "dssad_type": vehicle_info.dssad_type,
                "hardware_model": vehicle_info.hardware_model,
                "hardware_serial": vehicle_info.hardware_serial,
                "software_version": vehicle_info.software_version,
                "export_time": datetime.now().isoformat()
            }, f, ensure_ascii=False, indent=2)

        # 按事件导出
        for event in events:
            event_dir = os.path.join(export_dir, event.id)
            ensure_dir(event_dir)
            meta = {
                "id": event.id,
                "name": event.name,
                "event_type": event.event_type.value,
                "event_type_code": event.event_type_code,
                "is_locked": event.is_locked,
                "start_time": event.start_time.isoformat() if event.start_time else None,
                "end_time": event.end_time.isoformat() if event.end_time else None,
                "latitude": event.latitude,
                "longitude": event.longitude,
                "integrity_check": event.integrity_check,
                "is_tampered": event.is_tampered,
                "video_files": event.video_files,
                "non_video_files": event.non_video_files,
                "camera_channels": event.camera_channels
            }
            with open(os.path.join(event_dir, "metadata.json"), "w", encoding="utf-8") as f:
                json.dump(meta, f, ensure_ascii=False, indent=2)

            # 数据点序列
            if event.data_points:
                with open(os.path.join(event_dir, "data_points.json"), "w", encoding="utf-8") as f:
                    json.dump([dp.values for dp in event.data_points], f, ensure_ascii=False, indent=2)

            # 复制视频文件（模拟）
            for vf in event.video_files:
                if os.path.exists(vf):
                    shutil.copy2(vf, event_dir)

        return export_dir

    def convert_to_excel(self, events: List[Event]) -> str:
        """将事件转换为 Excel 格式，返回文件路径"""
        wb = Workbook()
        ws = wb.active
        ws.title = "事件汇总"

        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = [
            "事件ID", "事件名称", "事件类型", "事件类型编码", "是否锁定",
            "数据完整性", "是否被篡改", "经度", "纬度", "事件起点时间", "事件终点时间"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row, event in enumerate(events, 2):
            ws.cell(row=row, column=1, value=event.id)
            ws.cell(row=row, column=2, value=event.name)
            ws.cell(row=row, column=3, value=event.event_type.value)
            ws.cell(row=row, column=4, value=event.event_type_code)
            ws.cell(row=row, column=5, value="是" if event.is_locked else "否")
            ws.cell(row=row, column=6, value=event.integrity_check)
            ws.cell(row=row, column=7, value="是" if event.is_tampered else "否")
            ws.cell(row=row, column=8, value=event.longitude)
            ws.cell(row=row, column=9, value=event.latitude)
            ws.cell(row=row, column=10, value=event.start_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3])
            ws.cell(row=row, column=11, value=event.end_time.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3] if event.end_time else "")

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 24

        excel_path = os.path.join(
            self.output_dir,
            f"DSSAD_EVENTS_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        wb.save(excel_path)
        return excel_path
